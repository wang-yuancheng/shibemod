import csv
import openpyxl
from typing import cast
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from model import classify_sentence

wb = openpyxl.Workbook()
ws = cast(Worksheet, wb.active)
ws.append(["true", "text", "pred", "conf(%)", "status"])

green  = PatternFill(patternType="solid", fgColor="C6EFCE")
orange = PatternFill(patternType="solid", fgColor="FFD966")
red    = PatternFill(patternType="solid", fgColor="F4CCCC")
threshold = 0.80

with open("discord-final.csv", encoding="utf-8") as f:
    reader = csv.reader(f)
    next(reader)
    for true_label, text, *_ in reader:
        pred, probs = classify_sentence(text)
        conf = probs[pred]
        if pred == int(true_label):
            status, fill = "correct", green
        elif conf < threshold:
            status, fill = "wrong low", orange
        else:
            status, fill = "wrong high", red

        ws.append([true_label, text, pred, round(conf * 100, 2), status])
        for cell in ws[ws.max_row]:
            cell.fill = fill

wb.save("classified_output.xlsx")
